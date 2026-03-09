import os
import pymysql
import pandas as pd
from flask import Flask, request, jsonify, send_from_directory, render_template, redirect
from flask_cors import CORS
from werkzeug.utils import secure_filename
from flaskext.mysql import MySQL
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ================= DATABASE =================
app.config['MYSQL_DATABASE_USER'] = 'root'
app.config['MYSQL_DATABASE_PASSWORD'] = ''
app.config['MYSQL_DATABASE_DB'] = 'asset_management'
app.config['MYSQL_DATABASE_HOST'] = 'localhost'

mysql = MySQL()
mysql.init_app(app)

# ================= User =================
@app.route("/users")
def users():

    conn = mysql.connect()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute("SELECT * FROM users")
    users = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("users.html", users=users)

@app.route("/users/delete/<int:id>")
def delete_user(id):

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM users WHERE user_id=%s",(id,))
    conn.commit()

    cursor.close()
    conn.close()

    return redirect("/users")

@app.route("/users/add")
def add_user_page():
    return render_template("add_users.html")

@app.route("/users/insert", methods=["POST"])
def insert_user():

    first = request.form["first_name"]
    last = request.form["last_name"]
    email = request.form["email"]
    username = request.form["username"]
    password = request.form["password"]
    role = request.form["role"]

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute("""
    INSERT INTO users
    (first_name,last_name,email,username,password,role)
    VALUES (%s,%s,%s,%s,%s,%s)
    """,(first,last,email,username,password,role))

    conn.commit()

    cursor.close()
    conn.close()

    return redirect("/users")

@app.route("/users/edit/<int:id>")
def edit_user(id):

    conn = mysql.connect()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute("SELECT * FROM users WHERE user_id=%s",(id,))
    user = cursor.fetchone()

    cursor.close()
    conn.close()

    return render_template("edit_users.html", user=user)

@app.route("/users/update/<int:id>", methods=["POST"])
def update_user(id):

    first = request.form["first_name"]
    last = request.form["last_name"]
    email = request.form["email"]
    username = request.form["username"]
    password = request.form["password"]
    role = request.form["role"]

    conn = mysql.connect()
    cursor = conn.cursor()

    if password != "":
        cursor.execute("""
        UPDATE users SET
        first_name=%s,
        last_name=%s,
        email=%s,
        username=%s,
        password=%s,
        role=%s
        WHERE user_id=%s
        """,(first,last,email,username,password,role,id))
    else:
        cursor.execute("""
        UPDATE users SET
        first_name=%s,
        last_name=%s,
        email=%s,
        username=%s,
        role=%s
        WHERE user_id=%s
        """,(first,last,email,username,role,id))

    conn.commit()

    cursor.close()
    conn.close()

    return redirect("/users")

@app.route("/users/search")
def search_users():

    keyword = request.args.get("keyword")

    conn = mysql.connect()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute("""
    SELECT * FROM users
    WHERE first_name LIKE %s
    OR last_name LIKE %s
    OR username LIKE %s
    """,('%'+keyword+'%','%'+keyword+'%','%'+keyword+'%'))

    users = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("users.html", users=users)

# ================= Asset Types =================
@app.route("/asset_types")
def asset_types():

    conn = mysql.connect()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute("SELECT * FROM asset_types")
    types = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("asset_types.html", types=types)

@app.route("/asset_types/add")
def add_type_page():
    return render_template("add_type.html")

@app.route("/asset_types/edit/<int:id>")
def edit_type(id):

    conn = mysql.connect()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute("SELECT * FROM asset_types WHERE type_id=%s",(id,))
    type = cursor.fetchone()

    cursor.close()
    conn.close()

    return render_template("edit_type.html", type=type)

@app.route("/asset_types/search")
def search_types():

    keyword = request.args.get("keyword")

    conn = mysql.connect()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute("""
    SELECT * FROM asset_types
    WHERE type_name LIKE %s
    """,('%'+keyword+'%',))

    types = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("asset_types.html", types=types)

@app.route("/asset_types/insert", methods=["POST"])
def insert_type():

    type_name = request.form["type_name"]

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute(
        "INSERT INTO asset_types(type_name) VALUES(%s)",
        (type_name,)
    )

    conn.commit()

    cursor.close()
    conn.close()

    return redirect("/asset_types")

@app.route("/asset_types/delete/<int:id>")
def delete_type(id):

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM asset_types WHERE type_id=%s",
        (id,)
    )

    conn.commit()

    cursor.close()
    conn.close()

    return redirect("/asset_types")

# ================= IMPORT ASSET =================
@app.route("/admin/import", methods=["POST"])
def import_assets():

    file = request.files["file"]

    wb = load_workbook(file)
    sheet = wb.active

    conn = mysql.connect()
    cursor = conn.cursor()

    for row in sheet.iter_rows(min_row=2, values_only=True):

        asset_code = row[0]
        asset_name = row[1]
        brand = row[2]
        location = row[3]
        status = row[4]

        cursor.execute("""
        INSERT INTO assets
        (asset_code,asset_name,brand,location,status)
        VALUES (%s,%s,%s,%s,%s)
        """,(asset_code,asset_name,brand,location,status))

    conn.commit()

    cursor.close()
    conn.close()

    return redirect("/admin")
# ================= ADMIN PAGE =================
@app.route("/admin")
def admin():

    conn = mysql.connect()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute("SELECT * FROM assets")
    assets = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("admin.html", assets=assets)


# ================= ADMIN ADD PAGE =================
@app.route("/admin/add")
def admin_add():

    conn = mysql.connect()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute("SELECT * FROM asset_types")
    types = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("add.html", types=types)


# ================= ADMIN INSERT =================
@app.route("/admin/insert", methods=["POST"])
def admin_insert():

    code = request.form["asset_code"]
    name = request.form["asset_name"]
    brand = request.form["brand"]
    location = request.form["location"]
    status = request.form["status"]
    type_id = request.form["type_id"]

    image = request.files.get("image")
    filename = ""

    if image and image.filename != "":
        filename = secure_filename(image.filename)
        image.save(os.path.join(UPLOAD_FOLDER, filename))

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO assets
        (asset_code,asset_name,type_id,brand,location,status,image)
        VALUES (%s,%s,%s,%s,%s,%s,%s)
    """,(code,name,type_id,brand,location,status,filename))

    conn.commit()

    cursor.close()
    conn.close()

    return redirect("/admin")


# ================= ADMIN DELETE =================
@app.route("/admin/delete/<int:id>")
def admin_delete(id):

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM assets WHERE asset_id=%s",(id,))
    conn.commit()

    cursor.close()
    conn.close()

    return redirect("/admin")


# ================= ADMIN EDIT =================
@app.route("/admin/edit/<int:id>")
def admin_edit(id):

    conn = mysql.connect()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute("SELECT * FROM assets WHERE asset_id=%s",(id,))
    asset = cursor.fetchone()

    cursor.execute("SELECT * FROM asset_types")
    types = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("edit.html", asset=asset, types=types)


# ================= ADMIN UPDATE =================
@app.route("/admin/update/<int:id>", methods=["POST"])
def admin_update(id):

    code = request.form["asset_code"]
    name = request.form["asset_name"]
    brand = request.form["brand"]
    location = request.form["location"]
    status = request.form["status"]
    type_id = request.form["type_id"]

    image = request.files.get("image")

    conn = mysql.connect()
    cursor = conn.cursor()

    if image and image.filename != "":

        filename = secure_filename(image.filename)
        image.save(os.path.join(UPLOAD_FOLDER, filename))

        cursor.execute("""
        UPDATE assets SET
        asset_code=%s,
        asset_name=%s,
        type_id=%s,
        brand=%s,
        location=%s,
        status=%s,
        image=%s
        WHERE asset_id=%s
        """,(code,name,type_id,brand,location,status,filename,id))

    else:

        cursor.execute("""
        UPDATE assets SET
        asset_code=%s,
        asset_name=%s,
        type_id=%s,
        brand=%s,
        location=%s,
        status=%s
        WHERE asset_id=%s
        """,(code,name,type_id,brand,location,status,id))

    conn.commit()

    cursor.close()
    conn.close()

    return redirect("/admin")


# ================= GET ASSETS =================
@app.route("/assets", methods=["GET"])
def get_assets():

    search = request.args.get("search")

    conn = mysql.connect()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    if search:

        cursor.execute("""
        SELECT a.*, t.type_name
        FROM assets a
        LEFT JOIN asset_types t ON a.type_id = t.type_id
        WHERE a.asset_name LIKE %s
        OR a.asset_code LIKE %s
        """,('%'+search+'%','%'+search+'%'))

    else:

        cursor.execute("""
        SELECT a.*, t.type_name
        FROM assets a
        LEFT JOIN asset_types t ON a.type_id = t.type_id
        """)

    data = cursor.fetchall()

    cursor.close()
    conn.close()

    return jsonify(data)


# ================= GET ASSET BY CODE =================
@app.route("/assets/code/<code>", methods=["GET"])
def get_asset_by_code(code):
    conn = mysql.connect()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute("""
    SELECT a.*, t.type_name
    FROM assets a
    LEFT JOIN asset_types t ON a.type_id = t.type_id
    WHERE a.asset_code=%s
    """,(code,))

    data = cursor.fetchone()

    cursor.close()
    conn.close()

    if data:
        return jsonify(data)
    else:
        return jsonify({"error":"not found"}),404


# ================= DELETE ASSET API =================
@app.route("/assets/<int:id>", methods=["DELETE"])
def delete_asset_api(id):

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM assets WHERE asset_id=%s",(id,))
    conn.commit()

    cursor.close()
    conn.close()

    return jsonify({"message":"deleted"})

# ================= ADD ASSET API =================
@app.route("/assets", methods=["POST"])
def add_asset_api():

    try:

        data = request.json

        conn = mysql.connect()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO assets
            (asset_code,asset_name,type_id,brand,location,description,status,image)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """,(
            data.get("asset_code"),
            data.get("asset_name"),
            data.get("type_id"),
            data.get("brand"),
            data.get("location"),
            data.get("description",""),
            data.get("status"),
            data.get("image","")
        ))

        conn.commit()

        cursor.close()
        conn.close()

        return jsonify({"message":"added"})

    except Exception as e:
        print("ERROR:",e)
        return jsonify({"error":str(e)}),500

# ================= UPDATE ASSET API =================
@app.route("/assets/<int:id>", methods=["PUT"])
def update_asset_api(id):

    data = request.json

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE assets SET
        asset_code=%s,
        asset_name=%s,
        type_id=%s,
        brand=%s,
        location=%s,
        description=%s,
        status=%s,
        image=%s
        WHERE asset_id=%s
    """,(
        data["asset_code"],
        data["asset_name"],
        data["type_id"],
        data["brand"],
        data["location"],
        data["description"],
        data["status"],
        data["image"],
        id
    ))

    conn.commit()

    cursor.close()
    conn.close()

    return jsonify({"message":"updated"})


# ================= GET TYPES =================
@app.route("/types", methods=["GET"])
def get_types():

    conn = mysql.connect()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute("SELECT * FROM asset_types")
    data = cursor.fetchall()

    cursor.close()
    conn.close()

    return jsonify(data)


# ================= DASHBOARD =================
@app.route("/dashboard", methods=["GET"])
def dashboard():

    conn = mysql.connect()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute("SELECT COUNT(*) total FROM assets")
    total = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) normal FROM assets WHERE status='ปกติ'")
    normal = cursor.fetchone()["normal"]

    cursor.execute("SELECT COUNT(*) repair FROM assets WHERE status='แจ้งซ่อม'")
    repair = cursor.fetchone()["repair"]

    cursor.execute("SELECT COUNT(*) disposed FROM assets WHERE status='จำหน่ายออก'")
    disposed = cursor.fetchone()["disposed"]

    cursor.close()
    conn.close()

    return jsonify({
        "total":total,
        "normal":normal,
        "repair":repair,
        "disposed":disposed
    })


# ================= UPLOAD IMAGE =================
@app.route("/upload", methods=["POST"])
def upload():

    file = request.files['file']
    filename = secure_filename(file.filename)

    file.save(os.path.join(UPLOAD_FOLDER, filename))

    return jsonify({"filename":filename})


# ================= SHOW IMAGE =================
@app.route("/uploads/<filename>")
def uploaded_file(filename):

    return send_from_directory(UPLOAD_FOLDER, filename)

# ================= register =================
@app.route("/register", methods=["POST"])
def register():

    data = request.json

    conn = mysql.connect()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO users
        (first_name,last_name,email,username,password,role)
        VALUES (%s,%s,%s,%s,%s,'user')
    """,(
        data["first_name"],
        data["last_name"],
        data["email"],
        data["username"],
        data["password"]
    ))

    conn.commit()

    cursor.close()
    conn.close()

    return jsonify({"message":"success"})
# ================= login =================
@app.route("/login", methods=["POST"])
def login():

    data = request.json

    conn = mysql.connect()
    cursor = conn.cursor(pymysql.cursors.DictCursor)

    cursor.execute("""
        SELECT user_id,first_name,last_name,email,username,role
        FROM users
        WHERE username=%s AND password=%s
    """,(data["username"], data["password"]))

    user = cursor.fetchone()

    cursor.close()
    conn.close()

    if user:
        return jsonify({"status":"success","user":user})
    else:
        return jsonify({"status":"error"}),401


# ================= START SERVER =================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)